from django.shortcuts import render, redirect
from .models import *
# Create your views here.
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from openpyxl import load_workbook, Workbook
import tempfile
import os
import json


def home(request):
    all_files = UploadedFile.objects.all().order_by('-created_at')
    if request.method == "POST":
        file = request.FILES.get("file")
        print(f'This is File: {file}')
        
        uploaded_file = UploadedFile(file=file)
        uploaded_file.save()
        
        print(f'Loadind File: {file}')
        
        wb = load_workbook(file)
        ws = wb.active
        print(f'Loading File.............')
        # Extract headers from the first row
        headers = [cell.value for cell in ws[1]]
        print(f'Extracting Headers.............')
        uploaded_file.headers = headers
        uploaded_file.save()
        print(f'Saving Headers.............')

        return redirect("header_checks", pk=uploaded_file.pk)

    context = {
        'all_files': all_files,
    }
    return render(request, "index.html", context)


def home2(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        print(file)
        
    return render(request,"file-upload/index.html")


def all_names(request):
    existing_object = muslimNames.objects.first()
    print(existing_object)
    if request.method == "POST":
        name = request.POST.get("name")
        gender = request.POST.get("gender")
        file = request.FILES.get("file_with_name")
        
        name.lower()
        if name:
            print(name, gender)
            # Update the existing object
            if gender == 'male':
                existing_object.males.append(name)
            elif gender == 'female':
                existing_object.females.append(name)
            existing_object.save()

            return redirect(request.path)

        elif file:
            print(file, gender)
            wb = load_workbook(file)
            sheet = wb.active

            names = []
            for row in sheet.iter_rows(min_row=2, max_col=1):
                name = row[0].value
                names.append(name)

        # Check if an object already exists

            if existing_object:
                # Update the existing object
                if gender == 'male':
                    existing_object.males = names
                elif gender == 'female':
                    existing_object.females = names
                existing_object.save()
            else:
                # Create a new object
                if gender == 'male':
                    muslimNames.objects.create(males=names)
                elif gender == 'female':
                    muslimNames.objects.create(females=names)

    return render(request, "all_names.html")


def check_name(request):
    existing_object = muslimNames.objects.first()

    name = request.POST.get('name')
    name = name.lower()
    # Assuming you have lists of female_names and male_names containing names
    female_names = existing_object.females  # List of female names
    male_names = existing_object.males  # List of male names

    if name in female_names:
        return HttpResponse('<div style="color: red" class="fw-bold">This name already exists in the female names list!</div>')
    elif name in male_names:
        return HttpResponse('<div style="color: red" class="fw-bold">This name already exists in the male names list!</div>')
    else:
        return HttpResponse('<div style="color: green" class="fw-bold">This name does not exist in either list!</div>')


def header_checks(request, pk):
    file = UploadedFile.objects.get(pk=pk)
    count = len(file.headers) if file.headers else 0
    existing_object = muslimNames.objects.first()
    female_names = existing_object.females  # List of female names
    male_names = existing_object.males  # List of male names
    
    if request.method == "POST":
        predefined = request.POST.getlist('predefined')
        file_header = request.POST.getlist('file_header')

        
        # Load the Excel file
        wb = load_workbook(file.file)
        ws = wb.active
        # Find the index of the "persongender" column
        gender_col_index = None
        for index, header in enumerate(file.headers):
            if header.lower() == 'persongender':
                gender_col_index = index
                break
        
        # Find the index of the "persongender" column
        personname_col_index = None
        for index, header in enumerate(file.headers):
            if header.lower() == 'personfirstname':
                personname_col_index = index
                break

        # Filter rows based on predefined filters and the "persongender" column
        filtered_rows = []
            # Start iterating from the second row
        for row in ws.iter_rows(min_row=2):
                filtered_rows.append([cell.value for cell in row])
                

        if 'musllim_arabic_name' in predefined:
            filtered_muslims = []
            for row in filtered_rows:
                personname = row[personname_col_index].strip().lower()
                if personname:
                    if personname.lower() in [obj.lower() for obj in male_names] or personname.lower() in [obj.lower() for obj in female_names]:
                        filtered_muslims.append(row)
            # Select columns based on file headers
            filtered_rows = filtered_muslims
        
        selected_columns = []
        for row in filtered_rows:
            selected_row = []
            for index, header in enumerate(file.headers):
                if header in file_header:
                    selected_row.append(row[index])
            selected_columns.append(selected_row)

        # Create a new workbook containing the selected columns
        new_wb = Workbook()
        new_ws = new_wb.active

        # Insert header names in the first row
        for col_index, header_name in enumerate(file_header, start=1):
            new_ws.cell(row=1, column=col_index, value=header_name)

        # Append selected data rows
        for row_index, row_data in enumerate(selected_columns, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                new_ws.cell(row=row_index, column=col_index,
                            value=cell_value)

        # Save the new Excel file to a temporary file
        temp_file_path = os.path.join(
            tempfile.gettempdir(), 'filtered_data.xlsx')
        new_wb.save(temp_file_path)

        # Serve the new Excel file as a downloadable response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="filtered_data.xlsx"'
        with open(temp_file_path, 'rb') as f:
            response.write(f.read())

        # Delete the temporary file
        os.remove(temp_file_path)

        file.total_downloads += 1
        file.save()

        return response
    context = {
        'file': file,
        'count': count,
    }
    return render(request, "header_checks.html", context)
